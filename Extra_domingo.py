from genericpath import exists
import os
import shutil
from openpyxl import Workbook, load_workbook
import time
import PyPDF2
import re

def main():

    Planilha_Extra = load_workbook("C:\\Users\\gabriel.fonseca\\OneDrive - Energisa\\Documentos\\Horas Extras Domingo.xlsx")
    Aba = Planilha_Extra.active
    lastrow = Aba.max_row
    linha = lastrow

    regiao = str(input("Qual a Região?\n")).upper()
    
    #Abrir Arquivo
    espelho = PyPDF2.PdfFileReader("C:\\Users\\gabriel.fonseca\\Downloads\\Espelhos SE 16.09 A 15.10.pdf")

    # Coletar número de páginas do arquivo
    NumPages = espelho.getNumPages()

    seguir_regiao = str(input("Deseja segui a busca na Regional {}?\n".format(regiao))).upper()

    while seguir_regiao == "SIM":

        Data = str(input("Qual a data? (Completa 'Dia/mes/ano')\n"))
        somahora = 0

        # Extrair texto do PDF
        for i in range(0, NumPages):
            PageObj = espelho.getPage(i)
            Text = PageObj.extractText() 
            
            if regiao in Text:

                if Data in Text:
                    a = Text.split("\n")
                    nome = str(a[6].split(": ")[1])
                    indicedata = a.index(Data)
                        
                    '''print(a[indicedata])
                    print(a[indicedata+1])
                    print(a[indicedata+2])
                    print(a[indicedata+3])
                    print(a[indicedata+4])
                    print(a[indicedata+5])
                    print(a[indicedata+6])
                    print(a[indicedata+7])'''

                    if a[indicedata+4] == "** D.S.R. **":
                        hcompletp = a[indicedata+5].split(":")
                        hora = int(hcompletp[0])*60
                        min = int(hcompletp[1])
                        somahora = somahora+hora+min
                        linha = linha+1

                        Aba["A{}".format(linha)] = nome    
                        Aba["B{}".format(linha)] = Data 
                        Aba["D{}".format(linha)] = a[indicedata+5]  
                        Aba["E{}".format(linha)] = '16.09 A 15.10'
                        Aba["F{}".format(linha)] = i+1        
                        
                    elif a[indicedata+6] == "** D.S.R. **":
                        hcompletp = a[indicedata+7].split(":")
                        hora = int(hcompletp[0])*60
                        min = int(hcompletp[1])
                        somahora = somahora+hora+min

                        linha = linha+1
                        
                        Aba["A{}".format(linha)] = nome    
                        Aba["B{}".format(linha)] = Data 
                        Aba["D{}".format(linha)] = a[indicedata+7]  
                        Aba["E{}".format(linha)] = '16.09 A 15.10'
                        Aba["F{}".format(linha)] = i+1   

            '''h = somahora/60
            hm = str(h).split(".")
            floatmin = "0.{}".format(hm[1])
            mint = round(float(floatmin)*60)
            if mint < 10:
            m = "0{}".format(str(mint))
            print("{}:{}".format(hm[0], m))'''

        seguir_regiao = str(input("Deseja segui a busca na Regional {}?\n".format(regiao))).upper()

    Planilha_Extra.save("C:\\Users\\gabriel.fonseca\\OneDrive - Energisa\\Documentos\\Horas Extras Domingo.xlsx")
    
main()